# -*- coding: utf-8 -*-
"""
LMS 영상 검수 핵심 로직 (설정값 기반 검수 + 변환)
"""
import subprocess, json, os, sys, urllib.request, logging
from fractions import Fraction
from datetime import datetime

logger = logging.getLogger("vfr_v2")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)


def is_url(path):
    return isinstance(path, str) and path.startswith(("http://", "https://"))


def _fetch_head_bytes(url, size=200000):
    """URL에서 앞부분만 Range 요청으로 읽어옴"""
    try:
        req = urllib.request.Request(
            url, headers={"Range": f"bytes=0-{size-1}", "User-Agent": "ffprobe/6.0"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.read()
    except Exception as e:
        logger.warning("_fetch_head_bytes 실패 [%s]: %s", url, e)
        return None

# Windows 전용 플래그 (Mac/Linux에서는 0으로 대체)
_NO_WINDOW = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

REPORT_FILENAME        = "LMS영상검수보고서.xlsx"          # lms_watch.py 호환용
REPORT_FILENAME_MANUAL = "LMS영상검수보고서_수동.xlsx"
REPORT_FILENAME_AUTO   = "LMS영상검수보고서_자동.xlsx"
REPORT_FILENAME_ALL    = "LMS영상검수보고서_통합.xlsx"

DEFAULT_SETTINGS = {
    # ── 검수 기준값 ──────────────────────────────
    "codec":       "h264",   # 권장 코덱
    "fps":         30,       # 기준 프레임레이트 (fps)
    "bitrate":     1000,     # 최소 비트레이트 (kbps)
    "min_width":   1280,     # 최소 해상도 가로 (px)
    "min_height":  720,      # 최소 해상도 세로 (px)
    "audio_sr":    48000,    # 오디오 샘플레이트 (Hz)
    "audio_ch":    2,        # 오디오 채널 수
    # ── 변환 설정값 ──────────────────────────────
    "conv_fps":      30,     # 출력 프레임레이트
    "conv_bitrate":  1000,   # 출력 비디오 비트레이트 (kbps)
    "conv_preset":   "medium",
    "conv_audio_br": 128,    # 출력 오디오 비트레이트 (kbps)
}

_SETTINGS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")

def load_settings():
    """settings.json 로드 → 없으면 DEFAULT_SETTINGS 반환"""
    try:
        with open(_SETTINGS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {**DEFAULT_SETTINGS, **data}
    except Exception:
        return dict(DEFAULT_SETTINGS)

def save_settings(new_settings):
    """변경된 설정을 settings.json에 저장"""
    merged = {**DEFAULT_SETTINGS, **new_settings}
    try:
        with open(_SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def get_bin(name):
    candidates = []
    if getattr(sys, "frozen", False):
        candidates.append(os.path.join(sys._MEIPASS, name + ".exe"))
        candidates.append(os.path.join(os.path.dirname(sys.executable), name + ".exe"))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), name + ".exe"))
    for p in candidates:
        if os.path.exists(p):
            return p
    return name


def parse_rate(rate_str):
    try:
        return float(Fraction(rate_str))
    except Exception:
        return 0.0


def check_dts(filepath):
    cmd = [
        get_bin("ffprobe"), "-v", "quiet",
        "-select_streams", "a:0",
        "-show_packets",
        "-show_entries", "packet=dts",
        "-of", "csv=p=0",
        "-read_intervals", "%+#300",
        filepath
    ]
    try:
        out = subprocess.check_output(cmd, stderr=subprocess.DEVNULL,
                                      creationflags=_NO_WINDOW)
        values = []
        for line in out.decode("utf-8", errors="ignore").splitlines():
            line = line.strip()
            if line and line != "N/A":
                try:
                    values.append(int(line))
                except ValueError:
                    pass
        for i in range(1, len(values)):
            if values[i] < values[i - 1]:
                return True
    except Exception as e:
        logger.warning("check_dts 실패 [%s]: %s", filepath, e)
    return False


# ── 검사 항목별 헬퍼 함수 ────────────────────────────────────

def _check_video_stream(st, s, result):
    """비디오 스트림 검사: 코덱 / 해상도 / VFR / time_base"""
    codec  = st.get("codec_name", "")
    width  = st.get("width", 0)
    height = st.get("height", 0)
    r_fps  = parse_rate(st.get("r_frame_rate",   "0/1"))
    a_fps  = parse_rate(st.get("avg_frame_rate", "0/1"))

    result["codec"]  = codec
    result["width"]  = width
    result["height"] = height
    result["fps"]    = round(a_fps, 2) if a_fps else round(r_fps, 2)

    if codec and codec != s["codec"]:
        if codec == "hevc":
            result["issues"].append("코덱 주의(H.265, 권장: H.264)")
        else:
            result["issues"].append(f"코덱 비권장({codec}, 권장: {s['codec']})")

    if width and height:
        if width < s["min_width"] or height < s["min_height"]:
            result["issues"].append(
                f"해상도 미달({width}×{height}, 최소: {s['min_width']}×{s['min_height']})")

    if a_fps == 0 or (r_fps > 0 and abs(r_fps - a_fps) > 0.1):
        result["vfr"] = True
        result["issues"].append("VFR(가변 프레임레이트)")
    elif a_fps > 0 and abs(a_fps - s["fps"]) > 0.5:
        result["issues"].append(
            f"프레임레이트 불일치({a_fps:.2f}fps, 권장: {s['fps']}fps)")

    if st.get("time_base", "") == "1/90000":
        result["dts_error"] = True
        result["issues"].append("time_base 오류(1/90000)")


def _check_audio_stream(st, s, result):
    """오디오 스트림 검사: 샘플레이트 / 채널 수"""
    sr = int(st.get("sample_rate", 0))
    ch = int(st.get("channels", 0))
    if sr != s["audio_sr"]:
        result["audio_ok"] = False
        result["issues"].append(f"오디오 샘플레이트 {sr}Hz(권장 {s['audio_sr']}Hz)")
    if ch < s["audio_ch"]:
        result["audio_ok"] = False
        result["issues"].append("오디오 모노(스테레오 필요)")


def _check_bitrate(result, s):
    """비트레이트 기준 체크"""
    if result["bitrate"] > 0 and result["bitrate"] < s["bitrate"]:
        result["issues"].append(
            f"비트레이트 부족({result['bitrate']}kbps, 최소: {s['bitrate']}kbps)")


def _check_faststart(filepath, result, is_url_src):
    """moov atom 위치 확인 (Faststart)"""
    if is_url_src:
        header = _fetch_head_bytes(filepath)
        if header is None:
            result["faststart"] = None   # 확인불가
            return
        header_data = header
    else:
        try:
            with open(filepath, "rb") as f:
                header_data = f.read(200000)
        except Exception as e:
            logger.warning("faststart 파일 읽기 오류 [%s]: %s", filepath, e)
            result["issues"].append("파일 읽기 오류")
            return

    moov = header_data.find(b"moov")
    mdat = header_data.find(b"mdat")
    result["faststart"] = moov != -1 and (mdat == -1 or moov < mdat)
    if not result["faststart"]:
        result["issues"].append("faststart 없음(moov atom 위치 오류)")


# ── 메인 검수 함수 ────────────────────────────────────────────

def check_file(filepath, settings=None):
    s = {**DEFAULT_SETTINGS, **(settings or {})}
    _is_url = is_url(filepath)

    filename = (filepath.split("?")[0].rstrip("/").split("/")[-1]
                if _is_url else os.path.basename(filepath))

    result = {
        "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "filename":    filename,
        "filepath":    filepath,
        "codec":       "",
        "width":       0,
        "height":      0,
        "fps":         0.0,
        "bitrate":     0,
        "vfr":         False,
        "audio_ok":    True,
        "faststart":   False,
        "dts_error":   False,
        "issues":      [],
    }

    # ── ffprobe 실행 ────────────────────────────────
    cmd = [
        get_bin("ffprobe"), "-v", "quiet",
        "-print_format", "json",
        "-show_streams", "-show_format",
        filepath
    ]
    try:
        out  = subprocess.check_output(cmd, stderr=subprocess.STDOUT,
                                       creationflags=_NO_WINDOW)
        data = json.loads(out)
    except subprocess.CalledProcessError as e:
        msg = e.output.decode("utf-8", errors="ignore")[:60]
        logger.error("ffprobe CalledProcessError [%s]: %s", filename, msg)
        result["issues"].append("ffprobe 오류: " + msg)
        return result
    except Exception as e:
        logger.error("ffprobe 오류 [%s]: %s", filename, e)
        result["issues"].append("ffprobe 오류: " + str(e)[:60])
        return result

    # ── 전체 비트레이트 ──────────────────────────────
    try:
        result["bitrate"] = int(data.get("format", {}).get("bit_rate", 0)) // 1000
    except Exception as e:
        logger.warning("비트레이트 파싱 오류 [%s]: %s", filename, e)

    # ── 스트림별 검사 ────────────────────────────────
    for st in data.get("streams", []):
        if st.get("codec_type") == "video":
            _check_video_stream(st, s, result)
        elif st.get("codec_type") == "audio":
            _check_audio_stream(st, s, result)

    # ── 비트레이트 / DTS / Faststart ─────────────────
    _check_bitrate(result, s)

    if not _is_url and check_dts(filepath):
        result["dts_error"] = True
        result["issues"].append("DTS 역전")

    _check_faststart(filepath, result, _is_url)

    logger.info("검수완료 [%s] 문제:%d건", filename, len(result["issues"]))
    return result


def save_txt_report(report_path, results):
    """선택된 결과를 TXT 보고서로 저장 (타팀 공유용)"""
    total         = len(results)
    problem_count = sum(1 for r in results if r["issues"])
    ok_count      = total - problem_count

    lines = []
    lines.append("=" * 62)
    lines.append("  LMS 영상 검수 보고서")
    lines.append(f"  생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 62)
    lines.append(f"  전체: {total}개  |  정상: {ok_count}개  |  문제: {problem_count}개")
    lines.append("=" * 62)
    lines.append("")

    if problem_count > 0:
        lines.append("[문제 파일]")
        lines.append("-" * 62)
        for r in results:
            if r["issues"]:
                res_str = f"{r.get('width',0)}×{r.get('height',0)}" if r.get('width') else "-"
                fps_str = f"{'CFR' if not r.get('vfr') else 'VFR'} {r.get('fps',0):.1f}fps"
                br_str  = f"{r.get('bitrate',0)}kbps" if r.get('bitrate') else "-"
                lines.append(f"  파일명    : {r['filename']}")
                lines.append(f"  코덱      : {r.get('codec','-').upper()}")
                lines.append(f"  해상도    : {res_str}")
                lines.append(f"  프레임레이트: {fps_str}")
                lines.append(f"  비트레이트 : {br_str}")
                lines.append(f"  Faststart : {'있음' if r.get('faststart') else '없음'}")
                lines.append(f"  오디오    : {'정상' if r.get('audio_ok') else '문제'}")
                lines.append(f"  타임스탬프: {'오류' if r.get('dts_error') else '정상'}")
                lines.append(f"  문제 항목 : {' / '.join(r['issues'])}")
                lines.append("")

    if ok_count > 0:
        lines.append("[정상 파일]")
        lines.append("-" * 62)
        for r in results:
            if not r["issues"]:
                lines.append(f"  {r['filename']}")
        lines.append("")

    lines.append("=" * 62)
    lines.append("[검수 기준]")
    lines.append(f"  · 코덱       : H.264")
    lines.append(f"  · 해상도     : 1280×720 이상")
    lines.append(f"  · 프레임레이트: CFR 30fps (VFR 시 재생 오류)")
    lines.append(f"  · 비트레이트  : 1000kbps 이상")
    lines.append(f"  · 오디오     : 48000Hz / 스테레오(2ch)")
    lines.append(f"  · Faststart  : moov atom 파일 앞에 위치 필요")
    lines.append(f"  · 타임스탬프  : time_base 정상, DTS 역전 없음")
    lines.append("=" * 62)

    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return True
    except Exception as e:
        print(f"TXT 저장 오류: {e}")
        return False


def convert_file(filepath, output_dir, settings=None):
    s = {**DEFAULT_SETTINGS, **(settings or {})}
    fname    = os.path.basename(filepath)
    out_name = os.path.splitext(fname)[0] + "_fix.mp4"
    out_path = os.path.join(output_dir, out_name)
    fps      = s["conv_fps"]
    gop      = fps * 2  # 키프레임 2초 간격

    cmd = [
        get_bin("ffmpeg"),
        "-fflags", "+genpts+igndts",
        "-i", filepath,
        "-vf", f"fps={fps},setpts=PTS-STARTPTS,setsar=1",
        "-c:v", "libx264",
        "-b:v", f"{s['conv_bitrate']}k",
        "-profile:v", "high", "-level:v", "4.1",
        "-pix_fmt", "yuv420p",
        "-g", str(gop), "-keyint_min", str(gop), "-sc_threshold", "0",
        "-preset:v", s["conv_preset"],
        "-bf", "2",
        "-c:a", "aac", "-b:a", f"{s['conv_audio_br']}k",
        "-ar", str(s["audio_sr"]), "-ac", str(s["audio_ch"]),
        "-af", "aresample=async=1",
        "-avoid_negative_ts", "make_zero",
        "-start_at_zero",
        "-max_muxing_queue_size", "1024",
        "-movflags", "+faststart",
        "-y", out_path
    ]
    proc = subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                          creationflags=_NO_WINDOW)
    if proc.returncode != 0:
        err_msg = proc.stderr.decode("utf-8", errors="ignore")[-200:].strip()
        raise RuntimeError(f"ffmpeg 변환 실패 [{fname}]: {err_msg}")
    return out_path


def _init_sheet_headers(ws):
    """새 시트에 헤더 스타일 적용, 데이터 시작 행 반환"""
    headers = ["감지일시", "파일명", "코덱", "해상도", "프레임레이트",
               "비트레이트", "Faststart", "오디오", "타임스탬프", "판정", "문제항목"]
    hfill  = PatternFill("solid", fgColor="1E2A3A")
    hfont  = Font(bold=True, color="FFFFFF", size=10)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="AAAAAA")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = hfill, hfont, halign, bdr
    ws.row_dimensions[1].height = 28
    for col, w in zip("ABCDEFGHIJK", [18, 32, 10, 14, 14, 14, 12, 10, 14, 10, 50]):
        ws.column_dimensions[col].width = w
    return 2


def save_excel(report_path, results, sheet_name="검수결과"):
    """
    report_path : 저장할 Excel 파일 경로
    results     : 결과 딕셔너리 리스트
    sheet_name  : 저장할 시트명 (기본 "검수결과")
                  ex) "수동검수" / "자동감시" / "통합"
    """
    if not EXCEL_OK:
        return False
    try:
        # 기존 파일 로드 또는 새 파일 생성
        if os.path.exists(report_path):
            try:
                wb = openpyxl.load_workbook(report_path)
            except Exception:
                os.rename(report_path, report_path + ".bak")
                wb = None
        else:
            wb = None

        if wb is None:
            # 새 워크북: 기본 시트를 sheet_name으로 설정
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            start_row = _init_sheet_headers(ws)
        else:
            if sheet_name in wb.sheetnames:
                # 기존 시트에 이어서 추가
                ws = wb[sheet_name]
                start_row = ws.max_row + 1
            else:
                # 새 시트 생성
                ws = wb.create_sheet(sheet_name)
                start_row = _init_sheet_headers(ws)

        ok_fill  = PatternFill("solid", fgColor="E8F5E9")
        err_fill = PatternFill("solid", fgColor="FFEBEE")
        thin = Side(style="thin", color="AAAAAA")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r in results:
            is_ok = len(r["issues"]) == 0
            fill  = ok_fill if is_ok else err_fill
            res_str = f"{r.get('width',0)}×{r.get('height',0)}" if r.get('width') else "-"
            fps_str = f"{'CFR' if not r.get('vfr') else 'VFR'} {r.get('fps',0):.1f}fps"
            br_str  = f"{r.get('bitrate',0)}kbps" if r.get('bitrate') else "-"
            row_data = [
                r.get("detected_at", ""),
                r["filename"],
                r.get("codec", "-").upper(),
                res_str, fps_str, br_str,
                "✅ 있음" if r.get("faststart") else "❌ 없음",
                "✅ 정상" if r.get("audio_ok") else "❌ 문제",
                "✅ 정상" if not r.get("dts_error") else "❌ 오류",
                "정상" if is_ok else "문제",
                " / ".join(r["issues"]) if r["issues"] else "-"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=col, value=val)
                cell.fill      = fill
                cell.border    = bdr
                cell.font      = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="center" if col in (3, 4, 5, 6, 7, 8, 9, 10) else "left",
                    vertical="center", wrap_text=True
                )
            ws.row_dimensions[start_row].height = 20
            start_row += 1

        ws.auto_filter.ref = ws.dimensions
        wb.save(report_path)
        return True
    except Exception as e:
        print(f"Excel 저장 오류: {e}")
        return False
